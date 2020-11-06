import openpyxl,datetime
import matplotlib.pyplot as plt
import numpy as np
Loan=[] #贷款记录 [[编号，时间，金额，时长，利息],[。。。。],[。。。。]。。。。]
Income=[] #收入列表,365天的记录 [[日期 从一月一日到12月31日],[每天在外资金]，[每天每笔资金利息和]，[每天利息累计]]
fn='test.xlsx'
wb=openpyxl.load_workbook(fn,data_only=True)
allsheets=wb.get_sheet_names()
#print("所有工作表=",allsheets)
ws=wb.get_sheet_by_name(allsheets[0])
print("当前工作表=",ws.title)
print("当前工作表行数：",ws.max_row)
print("当前工作表列数：",ws.max_column)
for i in range(0,ws.max_row):
    L1=[]
    for cell in list(ws.rows)[i]:
        L1.append(cell.value)
    #print(L1)
    Loan.append(L1)
#print(Loan)
print("--------------------------------")
#初始化每日营收表Income[]，365项，每项内容：1：日期（从年初开始），2：当天占用资金，3：当天收益，4：累计收益（从年初开始）
Income=[]
incomeday=[]
Fd=datetime.date(2020,1,1)
for i in range(1,366):
    Cyear=2020
    Cmonth=Fd.fromordinal(i).month
    Cday=Fd.fromordinal(i).day
    Cdate=datetime.date(Cyear,Cmonth,Cday)
    Ccapital=0
    Cinterest=0
    Ctotal=0
    incomeday=[Cdate,Ccapital,Cinterest,Ctotal]
    Income.append(incomeday)
#print(Income)
#从Loan表中读取业务，更新Income[] 表
for i in range(1,len(Loan)): #读每一笔业务
    print(Loan[i])
    Ldate=Loan[i][1]        #该笔业务的垫款日期
    a=abs(Ldate.date()-Fd)    #一年中第几天，定位Income[]的位置
    Ldays=a.days            #成int
    print(Ldays,end=",  ")             #年中第几天
    if Ldays+Loan[i][3]>365:
        Zdays=365
    else:
        Zdays=Ldays+Loan[i][3]
    print(Zdays)
    for j in range(Ldays,Zdays):
        Income[j][1]+=Loan[i][2]
        Income[j][2]+=Loan[i][2]*Loan[i][4]
        #print(Income[j])
sumdays=0
for i in range(0,len(Income)):
    sumdays+=Income[i][2]
    #print(sumdays,end=" ")
    Income[i][3]=sumdays
    #print(Income[i])

arrIncome=np.array(Income)
#print(arrIncome)

print("--------------------")
'''
Dates,Dcapitals,Dinterests,Dtotals=[],[],[],[]
for row in Income:
    Date=row[0]
    Dates.append(Date)
    Dcapital=int(row[1])
    Dcapitals.append(Dcapital)
    Dinterest=int(row[2])
    Dinterests.append(Dinterest)
    Dtotal=int(row[3])
    Dtotals.append(Dtotal)
fig=plt.figure(dpi=128,figsize=(10,6))
plt.plot(Dates,Dcapitals,c='red')
plt.plot(Dates,Dtotals,c='blue')
plt.show()
'''
fig=plt.figure(dpi=128,figsize=(11,6))
#plt.plot(arrIncome[...,0],arrIncome[...,1],c='red')
plt.subplot(2,1,1)
plt.plot(arrIncome[...,0],arrIncome[...,2],c='blue')
plt.title("每日资金占用",fontproperties="Microsoft YaHei")

plt.subplot(2,1,2)
plt.plot(arrIncome[...,0],arrIncome[...,3],c='green')
plt.title("利润趋势图",fontproperties="Microsoft YaHei")

plt.show()